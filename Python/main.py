import sdmx
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import io

BASE_ULR = "https://api.imf.org/external/sdmx/3.0/data/dataflow/IMF.RES/WEO/6.0.0/"

IMF_DATA = sdmx.Client("IMF_DATA")
data_msg = IMF_DATA.data("WEO", key="USA+BRA+DEU+ZAF+IND.PCPIPCH+NGDP_RPCH", params={"startPeriod": 2015, "endPeriod": 2024})

weo_series  = sdmx.to_pandas(data_msg)
weo_df = weo_series.reset_index()


weo_df = weo_df.drop(columns=['LATEST_ACTUAL_ANNUAL_DATA', 'OVERLAP', 'METHODOLOGY_NOTES', 'METHODOLOGY'])

weo_df.columns = weo_df.columns.str.lower()

df_wide = weo_df.pivot_table(
    index=['country', 'time_period'],
    columns='indicator',
    values='value'
).reset_index()

country_name_map = {
    'USA': 'United States',
    'BRA': 'Brazil',
    'DEU': 'Germany',
    'ZAF': 'South Africa',
    'IND': 'India'
}

df_wide['country'] = df_wide['country'].replace(country_name_map)
df_wide = df_wide.rename(columns={
    'NGDP_RPCH': 'GDP_Growth',
    'PCPIPCH': 'Inflation',
    'time_period': 'Year'
})

df_wide['Year'] = df_wide['Year'].astype(int)

summary = df_wide.groupby('country')[['GDP_Growth', 'Inflation']].mean().round(2)
summary = summary.reset_index()
summary.columns = ['Country', 'Avg GDP Growth (%)', 'Avg Inflation (%)']

plt.figure(figsize=(12, 6))
for country in df_wide['country'].unique():
    country_data = df_wide[df_wide['country'] == country]
    plt.plot(country_data['Year'], country_data['GDP_Growth'], 
             label=country, marker='o')

plt.title('GDP Growth (2015-2024)')
plt.xlabel('Year')
plt.ylabel('GDP Growth (%)')
plt.legend()
plt.grid(True)
gdp_chart = plt.gcf()

df_2024 = df_wide[df_wide['Year'] == 2024]

df_2024 = df_2024.sort_values('Inflation', ascending=False)

plt.figure(figsize=(10, 6))
bars = plt.bar(
    df_2024['country'], 
    df_2024['Inflation'], 
    color=['#1f77b4']
)

for bar in bars:
    height = bar.get_height()
    plt.text(
        bar.get_x() + bar.get_width() / 2, 
        height, 
        f'{height:.1f}%', 
        ha='center', 
        va='bottom'
    )

plt.title('Inflation Rates in 2024 by Country', fontsize=14)
plt.xlabel('Country', fontsize=12)
plt.ylabel('Inflation (%)', fontsize=12)
plt.grid(axis='y', linestyle='--', alpha=0.7)
plt.tight_layout()
inflation_chart = plt.gcf()

def export_to_excel(summary_df, gdp_chart, inflation_chart, filename='IMF_analysis.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = "Economic Indicators"
    
    # Write summary table headers
    for col_num, header in enumerate(summary_df.columns, 1):
        ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)
    
    # Write summary data
    for row_num, row in enumerate(summary_df.values, 2):
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Add GDP chart
    gdp_buffer = io.BytesIO()
    gdp_chart.savefig(gdp_buffer, format='png', dpi=100, bbox_inches='tight')
    gdp_buffer.seek(0)
    img_gdp = Image(gdp_buffer)
    img_gdp.width, img_gdp.height = 600, 400
    ws.add_image(img_gdp, 'A' + str(len(summary_df) + 3))
    
    # Add Inflation chart
    infl_buffer = io.BytesIO()
    inflation_chart.savefig(infl_buffer, format='png', dpi=100, bbox_inches='tight')
    infl_buffer.seek(0)
    img_infl = Image(infl_buffer)
    img_infl.width, img_infl.height = 600, 400
    ws.add_image(img_infl, 'A' + str(len(summary_df) + 25))
    
    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    wb.save(filename)
    print(f"Exported to {filename}")

# Export to Excel
export_to_excel(summary, gdp_chart, inflation_chart)